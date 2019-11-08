#! python3
import openpyxl

wb = openpyxl.load_workbook('raw_youth.xlsx')
# print(wb.sheetnames)
sheet = wb['Form Responses 1']
# print(sheet['A3'].value)
# print(sheet.cell(row=2, column=4).value)

# print(sheet.max_row)
# print(sheet.max_column)


# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#         print('--- END OF ROW ---')


# sheet = wb.active
for cellObj in sheet.columns[1]:
    print(cellObj.value)
